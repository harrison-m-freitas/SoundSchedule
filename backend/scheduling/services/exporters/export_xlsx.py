from __future__ import annotations

import calendar as pycal
from datetime import time
from io import BytesIO
from typing import Iterable, List, Tuple

from django.conf import settings
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from scheduling.domain.models import Service
from scheduling.domain.repositories import ServiceRepository

# =========================
# Utilidades
# =========================

def _parse_time(s: str | time) -> time:
    """Padroniza a entrada como time."""
    if isinstance(s, time):
        return s
    hh, mm = str(s).split(":")
    return time(int(hh), int(mm))

def _autosize_columns(ws, max_width: int = 60):
    """Ajusta a largura das colunas com base no conteúdo.

    Args:
        ws (_type_): A planilha do Excel.
        max_width (int, optional): A largura máxima da coluna. Defaults to 60.
    """
    for col_idx, column_cells in enumerate(ws.columns, start=1):
        length = 0
        for cell in column_cells:
            v = "" if cell.value is None else str(cell.value)
            length = max(length, len(v))
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(10, length + 2), max_width)


def _header(ws, labels: Iterable[str]):
    """Escreve o cabeçalho da planilha.

    Args:
        ws (_type_): A planilha do Excel.
        labels (Iterable[str]): Os rótulos das colunas.
    """
    ws.append(list(labels))
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"

# =========================
# Export principal
# =========================

def export_schedule_xlsx(year: int, month: int, out_path: str) -> str:
    """Exporta a programação para um arquivo XLSX.

    Args:
        year (int): O ano da programação.
        month (int): O mês da programação.
        out_path (str): O caminho de saída do arquivo XLSX.

    Returns:
        str: O caminho do arquivo XLSX gerado.
    """
    wb = Workbook()

    ws = wb.active
    ws.title = f"Schedule ({year}-{month:02d})"
    _header(ws, ["Data", "Hora", "Tipo", "Rótulo", "Membros"])

    services: List[Service] = list(ServiceRepository.month_services(year, month))
    for s in services:
        names = ", ".join(a.member.__str__() for a in s.assignments.all())
        c_data = ws.cell(row=ws.max_row + 1, column=1, value=s.date)
        c_time = ws.cell(row=ws.max_row,     column=2, value=s.time)
        ws.cell(row=ws.max_row, column=3, value=s.type)
        ws.cell(row=ws.max_row, column=4, value=s.label or "")
        ws.cell(row=ws.max_row, column=5, value=names)

        c_data.number_format = "DD/MM/YYYY"
        c_time.number_format = "HH:MM"
        c_data.alignment = Alignment(horizontal="center")
        c_time.alignment = Alignment(horizontal="center")
    _autosize_columns(ws)

    ws2 = wb.create_sheet(title="Cultos (Resumo)")
    month_name = pycal.month_name[month] # How change this to Language Pt-BR?
    _header(ws2, [month_name, "Manhã", "Noite"])

    morning = _parse_time(settings.DEFAULT_MORNING_TIME)
    evening = _parse_time(settings.DEFAULT_EVENING_TIME)

    cal = pycal.Calendar(firstweekday=0)
    sundays = [d for d in cal.itermonthdates(year, month) if d.month == month and d.weekday() == 6]

    svc_map = {}
    for s in services:
        if s.type != "Culto":
            continue
        svc_map[(s.date, s.time)] = ", ".join(a.member.__str__() for a in s.assignments.all())

    for d in sorted(sundays):
        m = svc_map.get((d, morning), "")
        e = svc_map.get((d, evening), "")
        c_data = ws2.cell(row=ws2.max_row + 1, column=1, value=d)
        ws2.cell(row=ws2.max_row, column=2, value=m)
        ws2.cell(row=ws2.max_row, column=3, value=e)
        c_data.number_format = "DD/MM/YYYY"

    _autosize_columns(ws2)

    wb.save(out_path)
    return out_path
